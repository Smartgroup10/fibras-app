"""
Script de inicialización de la base de datos.
Crea las tablas, el usuario admin y carga los datos del Excel.
"""
import os
import sys
from pathlib import Path

# Añadir el directorio raíz al path
sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import load_workbook

# Ruta al Excel
EXCEL_PATH = r"C:\Users\STEFANO.YEPEZ\Downloads\fibras2.xlsx"

def init_tables():
    """Crear las tablas en la base de datos."""
    from app.database import engine, Base
    from app.models import User, Linea

    print("Creando tablas...")
    Base.metadata.create_all(bind=engine)
    print("Tablas creadas.")

def create_admin_user():
    """Crear el usuario administrador por defecto."""
    from app.database import SessionLocal
    from app.models import User
    from app.auth import get_password_hash

    db = SessionLocal()
    try:
        # Verificar si ya existe
        existing = db.query(User).filter(User.username == "admin").first()
        if existing:
            print("El usuario 'admin' ya existe.")
            return

        admin = User(
            username="admin",
            email="admin@fibras.local",
            password_hash=get_password_hash("admin123"),
            role="admin",
            is_active=True
        )
        db.add(admin)
        db.commit()
        print("Usuario admin creado (admin / admin123)")
    finally:
        db.close()

def load_excel_data():
    """Cargar los datos del Excel en la base de datos."""
    from app.database import SessionLocal
    from app.models import Linea

    if not os.path.exists(EXCEL_PATH):
        print(f"Archivo Excel no encontrado: {EXCEL_PATH}")
        return

    db = SessionLocal()
    try:
        # Verificar si ya hay datos
        count = db.query(Linea).count()
        if count > 0:
            print(f"Ya existen {count} líneas en la base de datos. Saltando importación.")
            return

        print(f"Cargando datos desde {EXCEL_PATH}...")
        wb = load_workbook(filename=EXCEL_PATH)
        ws = wb.active

        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            numero_linea = str(row[0] or '').strip()
            if not numero_linea:
                continue

            linea = Linea(
                numero_linea=numero_linea,
                servicio_conectividad=str(row[1] or '').strip(),
                sede=str(row[2] or '').strip(),
                tipo_conectividad=str(row[3] or '').strip(),
                velocidad_conexion=str(row[4] or '').strip(),
                tipo_ip=str(row[5] or '').strip(),
                tipo_mantenimiento=str(row[6] or '').strip()
            )
            db.add(linea)
            imported += 1

        db.commit()
        print(f"Importadas {imported} líneas.")
    except Exception as e:
        print(f"Error cargando datos: {e}")
        db.rollback()
    finally:
        db.close()

if __name__ == "__main__":
    print("=" * 50)
    print("INICIALIZACIÓN DE BASE DE DATOS (SQLite)")
    print("=" * 50)

    print("\n1. Creando tablas...")
    init_tables()

    print("\n2. Creando usuario admin...")
    create_admin_user()

    print("\n3. Cargando datos del Excel...")
    load_excel_data()

    print("\n" + "=" * 50)
    print("INICIALIZACIÓN COMPLETADA")
    print("=" * 50)
    print("\nPuedes iniciar la aplicación con:")
    print("  cd fibras_app")
    print("  python -m uvicorn app.main:app --reload")
    print("\nAccede a: http://localhost:8000")
    print("Login: admin / admin123")
