"""
Script de migración: Reinicia la BD con nuevos campos y carga ambos proveedores.
"""
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import load_workbook

# Rutas a los Excel
MOVISTAR_PATH = r"C:\Users\STEFANO.YEPEZ\Downloads\fibras2.xlsx"
LCR_PATH = r"C:\Users\STEFANO.YEPEZ\Downloads\lcr.xlsx"
DB_PATH = r"C:\Users\STEFANO.YEPEZ\fibras_app\fibras.db"

def reset_database():
    """Eliminar y recrear la base de datos."""
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        print(f"Base de datos eliminada: {DB_PATH}")

    from app.database import engine, Base
    from app.models import User, Linea

    print("Creando tablas con nuevos campos...")
    Base.metadata.create_all(bind=engine)
    print("Tablas creadas.")

def create_admin_user():
    """Crear el usuario administrador."""
    from app.database import SessionLocal
    from app.models import User
    from app.auth import get_password_hash

    db = SessionLocal()
    try:
        admin = User(
            username="admin",
            email="admin@fibras.local",
            password_hash=get_password_hash("admin123"),
            role="admin",
            is_active=True
        )
        db.add(admin)
        db.commit()
        print("Usuario admin creado (admin / admin123)")
    finally:
        db.close()

def load_movistar_data():
    """Cargar datos de Movistar (fibras2.xlsx)."""
    from app.database import SessionLocal
    from app.models import Linea

    if not os.path.exists(MOVISTAR_PATH):
        print(f"Archivo Movistar no encontrado: {MOVISTAR_PATH}")
        return 0

    db = SessionLocal()
    try:
        print(f"Cargando datos Movistar desde {MOVISTAR_PATH}...")
        wb = load_workbook(filename=MOVISTAR_PATH)
        ws = wb.active

        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            numero_linea = str(row[0] or '').strip()
            if not numero_linea:
                continue

            linea = Linea(
                numero_linea=numero_linea,
                proveedor="Movistar",
                servicio_conectividad=str(row[1] or '').strip(),
                sede=str(row[2] or '').strip(),
                tipo_conectividad=str(row[3] or '').strip(),
                velocidad_conexion=str(row[4] or '').strip(),
                tipo_ip=str(row[5] or '').strip(),
                tipo_mantenimiento=str(row[6] or '').strip()
            )
            db.add(linea)
            imported += 1

        db.commit()
        print(f"Movistar: {imported} líneas importadas.")
        return imported
    except Exception as e:
        print(f"Error cargando Movistar: {e}")
        db.rollback()
        return 0
    finally:
        db.close()

def load_lcr_data():
    """Cargar datos de LCR (lcr.xlsx)."""
    from app.database import SessionLocal
    from app.models import Linea

    if not os.path.exists(LCR_PATH):
        print(f"Archivo LCR no encontrado: {LCR_PATH}")
        return 0

    db = SessionLocal()
    try:
        print(f"Cargando datos LCR desde {LCR_PATH}...")
        wb = load_workbook(filename=LCR_PATH)
        ws = wb.active

        # Columnas LCR:
        # 0: Código
        # 1: Hitos
        # 2: Backup 4G
        # 3: Solicitud
        # 4: Solicitud Proveedor
        # 5: Estado LCR
        # 6: Estado Proveedor
        # 7: Cliente
        # 8: Instalación PTRO
        # 9: Ventana Activación
        # 10: Preparado Servicio
        # 11: Conexión Equipos
        # 12: Inicio Facturación
        # 13: IUA
        # 14: IDONT
        # 15: Direccion IP
        # 16: Nº Admin
        # 17: Dirección Instalación
        # 18: Provincia
        # 19: Velocidad

        imported = 0
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Usar Código LCR como número de línea (columna 0)
            codigo = row[0]
            if not codigo:
                continue

            numero_linea = f"LCR-{codigo}"
            iua = str(row[13] or '').strip()  # Guardar IUA para referencia

            # Extraer cliente (columna 7) - formato: [código] NOMBRE
            cliente_raw = str(row[7] or '').strip()
            # Limpiar formato [código] del nombre
            if cliente_raw.startswith('['):
                try:
                    cliente = cliente_raw.split(']')[1].strip()
                except:
                    cliente = cliente_raw
            else:
                cliente = cliente_raw

            # Dirección = Dirección Instalación + Provincia
            direccion = str(row[17] or '').strip()
            provincia = str(row[18] or '').strip()
            if provincia:
                sede = f"{direccion}, {provincia}"
            else:
                sede = direccion

            linea = Linea(
                numero_linea=numero_linea,
                proveedor="LCR",
                servicio_conectividad=f"Fibra LCR (IUA: {iua})" if iua else "Fibra LCR",
                sede=sede,
                tipo_conectividad="FIBRA",
                velocidad_conexion=str(row[19] or '').strip(),  # Velocidad
                tipo_ip="IP ESTATICA",  # LCR usa IPs estáticas
                tipo_mantenimiento=str(row[5] or '').strip(),  # Estado LCR
                cliente=cliente,
                direccion_ip=str(row[15] or '').strip()  # Dirección IP
            )
            db.add(linea)
            imported += 1

        db.commit()
        print(f"LCR: {imported} líneas importadas.")
        return imported
    except Exception as e:
        print(f"Error cargando LCR: {e}")
        import traceback
        traceback.print_exc()
        db.rollback()
        return 0
    finally:
        db.close()

if __name__ == "__main__":
    print("=" * 60)
    print("MIGRACIÓN: Nuevos campos + Carga Movistar + LCR")
    print("=" * 60)

    print("\n1. Reiniciando base de datos...")
    reset_database()

    print("\n2. Creando usuario admin...")
    create_admin_user()

    print("\n3. Cargando datos Movistar...")
    movistar_count = load_movistar_data()

    print("\n4. Cargando datos LCR...")
    lcr_count = load_lcr_data()

    print("\n" + "=" * 60)
    print("MIGRACIÓN COMPLETADA")
    print("=" * 60)
    print(f"\nResumen:")
    print(f"  - Movistar: {movistar_count} líneas")
    print(f"  - LCR: {lcr_count} líneas")
    print(f"  - Total: {movistar_count + lcr_count} líneas")
    print("\nPuedes iniciar la aplicación con:")
    print("  cd fibras_app")
    print("  python -m uvicorn app.main:app --reload")
