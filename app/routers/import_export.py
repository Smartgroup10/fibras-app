from fastapi import APIRouter, Depends, Request, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from openpyxl import load_workbook
import csv
import io
from ..database import get_db
from ..models import Linea, User
from ..dependencies import require_editor, get_current_user

router = APIRouter(tags=["import_export"])
templates = Jinja2Templates(directory="templates")

@router.get("/import", response_class=HTMLResponse)
async def import_page(
    request: Request,
    user: User = Depends(require_editor),
    db: Session = Depends(get_db)
):
    return templates.TemplateResponse("import.html", {
        "request": request,
        "user": user,
        "result": None
    })

@router.post("/import")
async def import_excel(
    request: Request,
    file: UploadFile = File(...),
    user: User = Depends(require_editor),
    db: Session = Depends(get_db)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        return templates.TemplateResponse("import.html", {
            "request": request,
            "user": user,
            "result": {"error": "El archivo debe ser Excel (.xlsx o .xls)"}
        })

    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents))
        ws = wb.active

        # Leer encabezados
        headers = [cell.value for cell in ws[1]]

        # Mapeo de columnas
        col_map = {}
        for idx, header in enumerate(headers):
            if header:
                h = header.lower().strip()
                if 'línea' in h or 'linea' in h:
                    col_map['numero_linea'] = idx
                elif 'servicio' in h:
                    col_map['servicio_conectividad'] = idx
                elif 'sede' in h:
                    col_map['sede'] = idx
                elif 'tipo' in h and 'conectividad' in h:
                    col_map['tipo_conectividad'] = idx
                elif 'velocidad' in h:
                    col_map['velocidad_conexion'] = idx
                elif 'tipo' in h and 'ip' in h:
                    col_map['tipo_ip'] = idx
                elif 'mantenimiento' in h:
                    col_map['tipo_mantenimiento'] = idx

        imported = 0
        updated = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                numero_linea = str(row[col_map.get('numero_linea', 0)] or '').strip()
                if not numero_linea:
                    continue

                linea_data = {
                    'numero_linea': numero_linea,
                    'servicio_conectividad': str(row[col_map.get('servicio_conectividad', 1)] or '').strip(),
                    'sede': str(row[col_map.get('sede', 2)] or '').strip(),
                    'tipo_conectividad': str(row[col_map.get('tipo_conectividad', 3)] or '').strip(),
                    'velocidad_conexion': str(row[col_map.get('velocidad_conexion', 4)] or '').strip(),
                    'tipo_ip': str(row[col_map.get('tipo_ip', 5)] or '').strip(),
                    'tipo_mantenimiento': str(row[col_map.get('tipo_mantenimiento', 6)] or '').strip()
                }

                existing = db.query(Linea).filter(Linea.numero_linea == numero_linea).first()
                if existing:
                    for key, value in linea_data.items():
                        setattr(existing, key, value)
                    updated += 1
                else:
                    linea = Linea(**linea_data)
                    db.add(linea)
                    imported += 1

            except Exception as e:
                errors.append(f"Fila {row_idx}: {str(e)}")

        db.commit()

        return templates.TemplateResponse("import.html", {
            "request": request,
            "user": user,
            "result": {
                "success": True,
                "imported": imported,
                "updated": updated,
                "errors": errors[:10]  # Mostrar solo los primeros 10 errores
            }
        })

    except Exception as e:
        return templates.TemplateResponse("import.html", {
            "request": request,
            "user": user,
            "result": {"error": f"Error procesando archivo: {str(e)}"}
        })

@router.get("/export/csv")
async def export_csv(
    request: Request,
    db: Session = Depends(get_db)
):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    lineas = db.query(Linea).all()

    output = io.StringIO()
    writer = csv.writer(output)

    # Encabezados
    writer.writerow([
        'Nº de línea',
        'Servicio de conectividad',
        'Sede',
        'Tipo de conectividad',
        'Velocidad de conexión',
        'Tipo de IP',
        'Tipo de mantenimiento'
    ])

    # Datos
    for linea in lineas:
        writer.writerow([
            linea.numero_linea,
            linea.servicio_conectividad,
            linea.sede,
            linea.tipo_conectividad,
            linea.velocidad_conexion,
            linea.tipo_ip,
            linea.tipo_mantenimiento
        ])

    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=lineas_export.csv"}
    )
